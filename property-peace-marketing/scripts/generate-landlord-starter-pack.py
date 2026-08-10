from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab import rl_config
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "public" / "downloads" / "landlord-starter-pack"
OUT.mkdir(parents=True, exist_ok=True)

rl_config.invariant = 1
FIXED_DATETIME = datetime(2026, 1, 1, 0, 0, 0)
FIXED_ZIP_TIME = (2026, 1, 1, 0, 0, 0)

NAVY = colors.HexColor("#061e35")
BLUE = colors.HexColor("#217eff")
GREEN = colors.HexColor("#15803d")
PALE_GREEN = colors.HexColor("#eef9f1")
PALE_BLUE = colors.HexColor("#f2f7fc")
SLATE = colors.HexColor("#405a70")
LIGHT = colors.HexColor("#d9e3ec")

styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name="PackTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=24, leading=29, textColor=NAVY, alignment=TA_LEFT, spaceAfter=10))
styles.add(ParagraphStyle(name="PackSubtitle", parent=styles["BodyText"], fontName="Helvetica", fontSize=10.5, leading=15, textColor=SLATE, spaceAfter=14))
styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=15, leading=19, textColor=NAVY, spaceBefore=12, spaceAfter=7))
styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=12, textColor=SLATE))
styles.add(ParagraphStyle(name="Tiny", parent=styles["BodyText"], fontName="Helvetica", fontSize=7.5, leading=10, textColor=SLATE))
styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10, textColor=NAVY))
styles.add(ParagraphStyle(name="CellBold", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=NAVY))
styles.add(ParagraphStyle(name="CenteredSmall", parent=styles["Small"], alignment=TA_CENTER))


def footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(LIGHT)
    canvas.line(0.55 * inch, 0.47 * inch, 7.95 * inch, 0.47 * inch)
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(SLATE)
    canvas.drawString(0.55 * inch, 0.29 * inch, "Property Peace Landlord Starter Pack • propertypeace.io/resources/starter-pack")
    canvas.drawRightString(7.95 * inch, 0.29 * inch, f"Page {doc.page}")
    canvas.restoreState()


def make_pdf(filename: str, title: str, subtitle: str, story):
    path = OUT / filename
    doc = SimpleDocTemplate(str(path), pagesize=letter, rightMargin=0.55 * inch, leftMargin=0.55 * inch, topMargin=0.55 * inch, bottomMargin=0.62 * inch, title=title, author="Property Peace Editorial Team")
    intro = [Paragraph("PROPERTY PEACE • LANDLORD STARTER PACK", ParagraphStyle(name=f"Eyebrow-{filename}", parent=styles["Small"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=GREEN, spaceAfter=8)), Paragraph(title, styles["PackTitle"]), Paragraph(subtitle, styles["PackSubtitle"])]
    doc.build(intro + story, onFirstPage=footer, onLaterPages=footer)
    return path


def deterministic_zip_info(filename: str) -> ZipInfo:
    info = ZipInfo(filename, FIXED_ZIP_TIME)
    info.compress_type = ZIP_DEFLATED
    info.create_system = 3
    info.external_attr = 0o100644 << 16
    return info


def normalize_zip_metadata(path: Path) -> None:
    """Repack ZIP-based formats with fixed entry ordering and timestamps."""
    with ZipFile(path, "r") as source:
        entries = [(name, source.read(name)) for name in sorted(source.namelist())]

    temporary = path.with_suffix(f"{path.suffix}.tmp")
    with ZipFile(temporary, "w", ZIP_DEFLATED) as target:
        for name, data in entries:
            target.writestr(deterministic_zip_info(name), data)
    temporary.replace(path)


def check_table(items, widths=(0.28 * inch, 4.12 * inch, 2.55 * inch), bottom_padding=14):
    data = [[Paragraph("□", styles["CellBold"]), Paragraph(item, styles["Cell"]), Paragraph("Notes / date / initials", styles["Tiny"])] for item in items]
    table = Table(data, colWidths=list(widths), repeatRows=0)
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.55, LIGHT),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, LIGHT),
        ("BACKGROUND", (0, 0), (0, -1), PALE_GREEN),
        ("BACKGROUND", (2, 0), (2, -1), PALE_BLUE),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), bottom_padding),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return table


def section(title, items):
    return [KeepTogether([Paragraph(title, styles["Section"]), check_table(items)])]


def blank_fields(labels):
    data = []
    for idx in range(0, len(labels), 2):
        row = []
        for label in labels[idx:idx + 2]:
            row.append(Paragraph(f"<b>{label}</b><br/><br/>________________________________", styles["Small"]))
        if len(row) == 1:
            row.append("")
        data.append(row)
    table = Table(data, colWidths=[3.45 * inch, 3.45 * inch])
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOX", (0, 0), (-1, -1), 0.55, LIGHT), ("INNERGRID", (0, 0), (-1, -1), 0.35, LIGHT), ("BACKGROUND", (0, 0), (-1, -1), colors.white), ("PADDING", (0, 0), (-1, -1), 8)]))
    return table


def signature_fields(labels):
    data = [[Paragraph(f"<b>{label}</b><br/>________________________________", styles["Small"]) for label in labels]]
    table = Table(data, colWidths=[3.45 * inch, 3.45 * inch])
    table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("BOX", (0, 0), (-1, -1), 0.55, LIGHT), ("INNERGRID", (0, 0), (-1, -1), 0.35, LIGHT), ("BACKGROUND", (0, 0), (-1, -1), colors.white), ("PADDING", (0, 0), (-1, -1), 6)]))
    return table


created = []

# 1. Move-in / move-out inspection
story = [blank_fields(["Property address", "Unit", "Inspection type (move-in / move-out)", "Inspection date", "Landlord / agent", "Tenant(s)"]), Spacer(1, 8)]
rooms = {
    "Entry, halls, and safety": ["Doors, locks, keys, handles, and weatherstripping", "Walls, ceilings, trim, floors, and stairs", "Smoke alarms and carbon-monoxide alarms; record test date", "Railings, exterior steps, lighting, mailbox, and house numbers"],
    "Kitchen": ["Cabinets, drawers, counters, backsplash, and sink", "Faucet, drain, disposal, and visible plumbing; check for leaks", "Range / oven, hood, refrigerator, dishwasher, and supplied appliances", "Outlets, GFCI protection, lights, windows, screens, and flooring"],
    "Bathrooms": ["Tub / shower, surround, caulk, grout, toilet, vanity, and mirror", "Faucets, drains, water pressure, exhaust fan, and visible plumbing", "Walls, ceiling, doors, hardware, outlets, lights, windows, and flooring", "Check for moisture, staining, mildew, loose fixtures, or slow drains"],
    "Bedrooms and living areas": ["Walls, ceilings, trim, doors, closets, and flooring", "Windows, screens, coverings, locks, outlets, switches, and lights", "Heating / cooling vents or equipment", "Photograph existing marks, wear, and damage from consistent angles"],
    "Utilities and exterior": ["Heating, cooling, thermostat, water heater, and filters", "Electrical panel, plumbing shutoffs, meters, and supplied equipment", "Basement, attic, garage, storage, balcony, yard, and common areas", "Record utility readings, remotes, parking permits, and access devices"],
}
for heading, items in rooms.items():
    story += section(heading, items)
story += [Paragraph("Photo and handoff record", styles["Section"]), check_table(["Use date-stamped photos or video and name files by room", "Record key, remote, parking-pass, and access-device counts", "Both parties receive or can access the completed condition record", "Record follow-up repairs separately; do not alter the signed condition record"], bottom_padding=6), Spacer(1, 10), Paragraph("Acknowledgment", styles["Section"]), Paragraph("This checklist records observed condition; it is not a waiver of rights or a substitute for state/local requirements.", styles["Small"]), signature_fields(["Landlord / agent signature and date", "Tenant signature and date"])]
created.append(make_pdf("move-in-move-out-inspection-checklist.pdf", "Move-In / Move-Out Inspection Checklist", "A room-by-room condition record for consistent inspections, photographs, keys, and follow-up. Adapt it to your property and current local requirements.", story))

# 2. Turnover checklist
turnover = {
    "Close out the prior tenancy": ["Confirm possession and collect all keys, remotes, permits, and access devices", "Complete and preserve the move-out condition record and photographs", "Document tenant property left behind and follow applicable notice/storage rules", "Record final utility readings and update access codes after lawful possession"],
    "Inspect and scope work": ["Separate ordinary wear from damage under current local rules", "Check life-safety devices, locks, railings, windows, electrical, plumbing, HVAC, and moisture", "Create one repair list with owner, priority, estimate, approval, and target date", "Photograph hidden or high-cost conditions before work begins"],
    "Repair, clean, and verify": ["Complete licensed or permit-required work through qualified providers", "Patch, paint, clean, replace filters, service appliances, and address pests as needed", "Test every supplied appliance, fixture, light, lock, alarm, and water shutoff", "Walk the finished unit against the repair list and retain invoices/warranties"],
    "Prepare the next tenancy": ["Set rent and availability using current property and market information", "Confirm advertising, screening, and occupancy criteria are lawful and consistently applied", "Prepare current lease/addenda and required disclosures; obtain local review when needed", "Schedule move-in funds, utilities, keys, orientation, and the new condition inspection"],
}
story = [blank_fields(["Property / unit", "Prior move-out date", "Target ready date", "Target move-in date", "Turnover owner", "Approved budget"]), Spacer(1, 8)]
for heading, items in turnover.items(): story += section(heading, items)
story += [Paragraph("Turnover cost and status log", styles["Section"]), Table([[Paragraph("Work item", styles["CellBold"]), Paragraph("Owner / vendor", styles["CellBold"]), Paragraph("Estimate", styles["CellBold"]), Paragraph("Actual", styles["CellBold"]), Paragraph("Status / date", styles["CellBold"])] ] + [["", "", "", "", ""] for _ in range(8)], colWidths=[2.15*inch,1.5*inch,0.85*inch,0.85*inch,1.6*inch], rowHeights=[0.3*inch]+[0.38*inch]*8, style=TableStyle([("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.4,LIGHT),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("PADDING",(0,0),(-1,-1),5)])), Spacer(1, 10), Paragraph("Educational checklist only. Deposit accounting, abandoned property, habitability, licensing, notices, disclosures, and rekeying rules vary by location.", styles["Small"])]
created.append(make_pdf("rental-turnover-checklist.pdf", "Rental Turnover Checklist", "A practical sequence from possession through a documented, ready-to-rent handoff—without pretending one checklist replaces local law or qualified trades.", story))

# 3. Preventive maintenance calendar
months = [
    ("January", "Freeze protection, heating performance, ice/water intrusion, emergency contacts"), ("February", "Interior moisture, bathroom ventilation, leaks, appliance hoses, tenant check-in"), ("March", "Roof/gutter visual check, drainage, sump pump, exterior trip hazards"), ("April", "Cooling service, filters, screens, exterior faucets, pest-prevention gaps"), ("May", "Decks/railings, landscaping clearances, irrigation, exterior lighting"), ("June", "Roof/attic heat and moisture, dryer vents, plumbing shutoffs, midyear records"), ("July", "Cooling performance, condensation drains, pools/common areas where applicable"), ("August", "Weather seals, windows, doors, fall vendor scheduling, tree clearances"), ("September", "Heating service, filters, chimneys where applicable, freeze planning"), ("October", "Gutters/downspouts, exterior water shutoff, leaves/drainage, lighting"), ("November", "Smoke/CO alarms, emergency heat, snow/ice supplies, vacancy checks"), ("December", "Year-end invoices/warranties, recurring issue review, next-year maintenance budget"),
]
story = [blank_fields(["Property / portfolio", "Calendar year", "Primary contact", "Emergency vendor / contact"]), Spacer(1, 8), Paragraph("Monthly baseline", styles["Section"]), check_table(["Review open work orders, recurring leaks/moisture, pest reports, and safety issues", "Confirm required inspections and alarm tests under current local rules", "Record work performed, vendor, invoice, warranty, photographs, and next due date", "Give lawful notice before entry except where an emergency rule applies"]), PageBreak()]
for month, focus in months:
    story += [KeepTogether([Paragraph(month, styles["Section"]), Table([[Paragraph(f"Suggested focus: {focus}", styles["Cell"]), Paragraph("Property-specific tasks / due dates / owner", styles["Tiny"])] , ["", ""]], colWidths=[3.2*inch,3.75*inch], rowHeights=[0.36*inch,0.58*inch], style=TableStyle([("BACKGROUND",(0,0),(0,0),PALE_GREEN),("BACKGROUND",(1,0),(1,0),PALE_BLUE),("GRID",(0,0),(-1,-1),0.4,LIGHT),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),6)]))]), Spacer(1, 4)]
story += [Spacer(1, 8), Paragraph("This calendar is a planning aid, not an inspection standard. Climate, building systems, manufacturer schedules, lease duties, and local codes should determine the final plan.", styles["Small"])]
created.append(make_pdf("preventive-maintenance-calendar.pdf", "Preventive Maintenance Calendar", "A twelve-month planning worksheet for small rental portfolios. Replace the suggested focus areas with the actual systems, climate, manufacturer guidance, and legal duties for each property.", story))

# 4. Fair-housing-safe pre-screening worksheet
story = [Paragraph("Use the same written process for every prospect", styles["Section"]), Paragraph("Pre-screening should identify objective rental-fit information without asking about protected traits. Publish your lawful criteria first, ask only what is necessary, offer reasonable-accommodation channels, and route every qualified prospect through the same next step.", styles["PackSubtitle"]), blank_fields(["Property / unit", "Advertised rent", "Available date", "Application link / next step", "Written criteria version/date", "Staff member"]), Spacer(1, 8), Paragraph("Suggested neutral questions", styles["Section"]), check_table(["What move-in date are you seeking?", "How many occupants would live in the home? Apply current lawful occupancy standards consistently.", "What lease term are you seeking?", "Do you have animals that would live in the home? Keep assistance-animal requests separate from pet screening.", "Can you meet the published income or other lawful financial criteria? Explain acceptable documentation and alternatives consistently.", "Can you meet the published move-in cost and payment schedule?", "Would you like to request a reasonable accommodation or an accessible application method?", "What is the best lawful way and time to contact you about the next step?"]), *section("Do not ask or record", ["Race, color, national origin, religion, sex, familial status, or disability details", "Questions about children, pregnancy, marital status, citizenship assumptions, or neighborhood preferences", "Medical diagnoses, medications, nature of a disability, or why an assistance animal is needed beyond permitted verification", "Different questions, standards, response times, or discouraging language for different prospects", "Protected information inferred from names, photos, accents, social media, or in-person impressions"]), Paragraph("Prospect log", styles["Section"]), Table([[Paragraph("Date/time", styles["CellBold"]), Paragraph("Prospect reference", styles["CellBold"]), Paragraph("Answers / objective notes", styles["CellBold"]), Paragraph("Next step / sent", styles["CellBold"])] ] + [["", "", "", ""] for _ in range(7)], colWidths=[1.0*inch,1.25*inch,3.5*inch,1.2*inch], rowHeights=[0.3*inch]+[0.46*inch]*7, style=TableStyle([("BACKGROUND",(0,0),(-1,0),NAVY),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),0.4,LIGHT),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),5)])), Spacer(1, 10), Paragraph("Educational worksheet only—not legal advice. Federal, state, and local fair-housing, source-of-income, screening, occupancy, privacy, and notice rules may add protections or requirements. Review current HUD guidance and obtain qualified local advice before adopting criteria.", styles["Small"])]
created.append(make_pdf("fair-housing-safe-pre-screening-worksheet.pdf", "Fair-Housing-Safe Pre-Screening Worksheet", "A neutral, repeatable starting point for early prospect conversations. It deliberately avoids collecting sensitive screening data and does not replace a lawful application, consent, or consumer-report process.", story))

# 5. Cash-flow workbook
wb = Workbook()
wb.properties.creator = "Property Peace Editorial Team"
wb.properties.lastModifiedBy = "Property Peace Editorial Team"
wb.properties.created = FIXED_DATETIME
wb.properties.modified = FIXED_DATETIME
ws = wb.active
assert ws is not None
ws.title = "Monthly Cash Flow"
navy_fill = PatternFill("solid", fgColor="061E35")
green_fill = PatternFill("solid", fgColor="15803D")
pale_fill = PatternFill("solid", fgColor="EEF9F1")
white_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D9E3EC")
for col, width in {"A":18,"B":15,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":13,"K":13,"L":13,"M":13,"N":15,"O":17}.items(): ws.column_dimensions[col].width = width
ws.merge_cells("A1:O1"); ws["A1"] = "Property Peace Rental Cash-Flow Workbook"; ws["A1"].fill = navy_fill; ws["A1"].font = Font(color="FFFFFF",bold=True,size=16); ws["A1"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:O2"); ws["A2"] = "Enter actual monthly amounts by property. Green cells are inputs; totals and annual figures calculate automatically."; ws["A2"].alignment = Alignment(horizontal="center"); ws["A2"].font = Font(color="405A70",italic=True)
headers = ["Category","Type"] + ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"] + ["Annual total"]
for c,h in enumerate(headers,1): cell=ws.cell(4,c,h); cell.fill=navy_fill; cell.font=white_font; cell.alignment=Alignment(horizontal="center",wrap_text=True); cell.border=Border(bottom=thin)
rows = [
("Gross rent scheduled","Income"),("Other income","Income"),("Vacancy / concessions","Reduction"),("Uncollected rent","Reduction"),("Operating income","Formula"),
("Repairs and maintenance","Expense"),("Utilities paid by owner","Expense"),("Insurance","Expense"),("Property taxes","Expense"),("Management / leasing","Expense"),("HOA / association","Expense"),("Legal / accounting","Expense"),("Licenses / inspections","Expense"),("Advertising","Expense"),("Other operating expense","Expense"),("Operating expenses","Formula"),("Net operating income (NOI)","Formula"),("Debt service","Financing"),("Capital improvements","Capital"),("Reserve contribution","Reserve"),("Cash flow before tax","Formula"),]
start=5
for i,(label,kind) in enumerate(rows,start):
    ws.cell(i,1,label); ws.cell(i,2,kind)
    for c in range(3,15):
        cell=ws.cell(i,c)
        if kind in {"Income","Reduction","Expense","Financing","Capital","Reserve"}: cell.fill=pale_fill
        cell.number_format='$#,##0.00;[Red]-$#,##0.00'; cell.border=Border(bottom=thin)
    ws.cell(i,15,f"=SUM(C{i}:N{i})"); ws.cell(i,15).number_format='$#,##0.00;[Red]-$#,##0.00'
row_by_label={label:i for i,(label,_) in enumerate(rows,start)}
for c in range(3,15):
    letter=get_column_letter(c)
    ws.cell(row_by_label["Operating income"],c,f"={letter}{row_by_label['Gross rent scheduled']}+{letter}{row_by_label['Other income']}-{letter}{row_by_label['Vacancy / concessions']}-{letter}{row_by_label['Uncollected rent']}")
    expense_rows=[row_by_label[x] for x in ["Repairs and maintenance","Utilities paid by owner","Insurance","Property taxes","Management / leasing","HOA / association","Legal / accounting","Licenses / inspections","Advertising","Other operating expense"]]
    ws.cell(row_by_label["Operating expenses"],c,"=SUM("+",".join(f"{letter}{r}" for r in expense_rows)+")")
    ws.cell(row_by_label["Net operating income (NOI)"],c,f"={letter}{row_by_label['Operating income']}-{letter}{row_by_label['Operating expenses']}")
    ws.cell(row_by_label["Cash flow before tax"],c,f"={letter}{row_by_label['Net operating income (NOI)']}-{letter}{row_by_label['Debt service']}-{letter}{row_by_label['Capital improvements']}-{letter}{row_by_label['Reserve contribution']}")
for label in ["Operating income","Operating expenses","Net operating income (NOI)","Cash flow before tax"]:
    r=row_by_label[label]
    for c in range(1,16): ws.cell(r,c).font=Font(bold=True,color="061E35"); ws.cell(r,c).fill=PatternFill("solid",fgColor="F2F7FC")
ws.freeze_panes="C5"; ws.auto_filter.ref=f"A4:O{start+len(rows)-1}"; ws.sheet_view.showGridLines=False
instructions=wb.create_sheet("Read Me")
instructions.column_dimensions["A"].width=22; instructions.column_dimensions["B"].width=95
notes=[("Purpose","A reusable monthly property cash-flow worksheet. Duplicate the Monthly Cash Flow sheet for each property."),("Inputs","Enter positive amounts in green cells. Vacancy, uncollected rent, expenses, debt service, improvements, and reserves are subtracted by formulas."),("NOI","Net operating income excludes debt service, capital improvements, reserves, depreciation, and income taxes."),("Do not mix","Keep refundable security deposits and owner contributions out of operating income. Track them in your accounting system according to applicable rules."),("Review","Reconcile entries to bank records, invoices, leases, and your accounting records. Have a qualified tax professional review tax treatment."),("Disclaimer","Educational template only. It is not accounting, tax, legal, lending, or investment advice.")]
instructions["A1"]="Rental Cash-Flow Workbook"; instructions["A1"].fill=navy_fill; instructions["A1"].font=Font(color="FFFFFF",bold=True,size=15); instructions.merge_cells("A1:B1")
for r,(a,b) in enumerate(notes,3): instructions.cell(r,1,a).font=Font(bold=True,color="15803D"); instructions.cell(r,2,b).alignment=Alignment(wrap_text=True,vertical="top"); instructions.row_dimensions[r].height=42
xlsx_path=OUT/"rental-property-cash-flow-workbook.xlsx"; wb.save(xlsx_path); normalize_zip_metadata(xlsx_path); created.append(xlsx_path)

# README and bundle
readme = OUT / "README.txt"
readme.write_text("""PROPERTY PEACE LANDLORD STARTER PACK\n\nIncluded files:\n- Move-In / Move-Out Inspection Checklist (PDF)\n- Rental Turnover Checklist (PDF)\n- Preventive Maintenance Calendar (PDF)\n- Fair-Housing-Safe Pre-Screening Worksheet (PDF)\n- Rental Property Cash-Flow Workbook (XLSX)\n\nThese resources are editable/printable planning aids and general educational information. They are not legal, tax, accounting, screening, safety, or compliance advice. Requirements vary by property and location. Verify current primary sources and consult qualified local professionals.\n\nSource: https://propertypeace.io/resources/starter-pack/\n""", encoding="utf-8")
created.append(readme)
zip_path = OUT / "property-peace-landlord-starter-pack.zip"
with ZipFile(zip_path, "w", ZIP_DEFLATED) as zf:
    for path in created:
        zf.writestr(deterministic_zip_info(path.name), path.read_bytes())

print(f"Generated {len(created)} pack files plus bundle: {zip_path}")
for path in created + [zip_path]:
    print(f"{path.name}\t{path.stat().st_size} bytes")
